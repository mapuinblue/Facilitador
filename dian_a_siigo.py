"""
DIAN a Siigo - Aplicación de Escritorio v3.1
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
from pathlib import Path
import re
from datetime import datetime
import os
import sys
import traceback


class ProcesadorContableDIAN:
    """Procesa archivos DIAN con detección automática de estructura"""
    
    def __init__(self):
        self.IVA_RATE = 0.19
        self.CUENTAS_COMPRAS = {
            'gasto': '14, 51, 61',
            'iva_descontable': '24080103'
        }
        self.CUENTAS_VENTAS = {
            'ingresos': '41',
            'iva_generado': '24080101',
            'iva_credito': '13050501'
        }
    
    def limpiar_numero(self, valor_str):
        """
        Limpia un número en formato de texto a float.
        Maneja formatos: 
        - 1234.56 (punto decimal)
        - 1.234,56 (punto miles, coma decimal - formato colombiano)
        - 1,234.56 (coma miles, punto decimal)
        - "1234,56" (coma decimal sin miles)
        """
        if pd.isna(valor_str) or valor_str == '' or valor_str is None:
            return 0.0
        
        # Convertir a string por si acaso
        valor_str = str(valor_str).strip()
        
        if valor_str == '' or valor_str == 'nan':
            return 0.0
        
        # Eliminar espacios
        valor_str = valor_str.replace(' ', '')
        
        # Detectar formato por la posición de comas y puntos
        tiene_punto = '.' in valor_str
        tiene_coma = ',' in valor_str
        
        try:
            if tiene_punto and tiene_coma:
                # Determinar cuál es el separador de miles y cuál el decimal
                pos_ultimo_punto = valor_str.rfind('.')
                pos_ultima_coma = valor_str.rfind(',')
                
                if pos_ultima_coma > pos_ultimo_punto:
                    # Formato: 1.234,56 (colombiano)
                    valor_str = valor_str.replace('.', '').replace(',', '.')
                else:
                    # Formato: 1,234.56 (inglés)
                    valor_str = valor_str.replace(',', '')
            elif tiene_coma and not tiene_punto:
                # Solo tiene coma - probablemente es decimal
                # Verificar si hay más de una coma
                if valor_str.count(',') == 1:
                    # Formato: 1234,56 - coma es decimal
                    valor_str = valor_str.replace(',', '.')
                else:
                    # Formato: 1,234,567 - comas son miles
                    valor_str = valor_str.replace(',', '')
            # Si solo tiene punto, dejar como está (formato estándar)
            
            return float(valor_str)
        except ValueError:
            print(f"⚠️ No se pudo convertir '{valor_str}' a número, usando 0")
            return 0.0
    
    def formato_pesos_display(self, valor):
        """Formatea número al estilo colombiano SOLO PARA MOSTRAR: 200.000,00"""
        try:
            if pd.isna(valor) or valor == 0 or valor == '':
                return "0,00"
            # Convertir a entero (redondeado al peso)
            valor_entero = int(round(float(valor), 0))
            # Formatear con punto para miles y coma para decimales
            return f"{valor_entero:,}".replace(",", ".") + ",00"
        except:
            return "0,00"
    
    def valor_numerico(self, valor):
        """Devuelve el valor redondeado al peso más cercano (sin decimales)"""
        try:
            if pd.isna(valor) or valor == '' or valor == 0:
                return None
            # Asegurar conversión correcta
            valor_float = float(valor)
            # Redondear al entero más cercano
            return int(round(valor_float, 0))
        except:
            return None
    
    def valor_numerico_base(self, valor):
        """Devuelve el valor de base redondeado al peso más cercano (sin decimales)"""
        try:
            if pd.isna(valor) or valor == '' or valor == 0:
                return None
            valor_float = float(valor)
            # Redondear al entero más cercano
            return int(round(valor_float, 0))
        except:
            return None
    
    def redondear_peso(self, valor):
        """Redondea al peso más cercano"""
        try:
            if pd.isna(valor) or valor == '':
                return 0
            valor_float = float(valor)
            return int(round(valor_float, 0))
        except:
            return 0
    
    def limpiar_nit(self, nit):
        """Limpia y formatea el NIT"""
        if pd.isna(nit):
            return ""
        nit_str = str(nit).strip().replace('.0', '').replace('.00', '')
        return re.sub(r'[^\d]', '', nit_str)
    
    def leer_archivo_dian(self, ruta_archivo):
        """
        Lee archivo DIAN con mejor detección de estructura:
        - Busca encabezados reales buscando patrones conocidos
        - Maneja diferentes formatos de archivo
        """
        extension = Path(ruta_archivo).suffix.lower()
        
        try:
            # Primero intentar leer sin saltar filas para inspeccionar
            if extension == '.csv':
                df_raw = pd.read_csv(ruta_archivo, encoding='utf-8-sig', 
                                   header=None, dtype=str, nrows=10)
            else:
                xl = pd.ExcelFile(ruta_archivo)
                hoja = xl.sheet_names[0]
                df_raw = pd.read_excel(ruta_archivo, sheet_name=hoja, 
                                     header=None, dtype=str, nrows=10)
            
            print("Primeras filas del archivo crudo:")
            for i in range(min(5, len(df_raw))):
                print(f"Fila {i}: {list(df_raw.iloc[i].dropna().head(15))}")
            
            # Buscar la fila que contiene encabezados clave
            header_row = None
            for idx, row in df_raw.iterrows():
                row_str = ' '.join([str(cell) for cell in row if pd.notna(cell)])
                if any(keyword in row_str.lower() for keyword in ['total', 'iva', 'nit', 'emisor', 'receptor']):
                    header_row = idx
                    print(f"Encontrado encabezado en fila {idx}")
                    break
            
            if header_row is None:
                # Usar la primera fila no vacía como encabezado
                for idx, row in df_raw.iterrows():
                    if not row.isnull().all():
                        header_row = idx
                        break
            
            print(f"Usando fila {header_row} como encabezado")
            
            # Leer el archivo con el encabezado encontrado
            if extension == '.csv':
                df = pd.read_csv(ruta_archivo, encoding='utf-8-sig', 
                               skiprows=header_row, header=0, dtype=str)
            else:
                df = pd.read_excel(ruta_archivo, sheet_name=hoja, 
                                 skiprows=header_row, header=0, dtype=str)
            
            # Limpiar nombres de columnas
            df.columns = [str(col).strip() for col in df.columns]
            
            # Mostrar columnas detectadas
            print(f"\nColumnas detectadas ({len(df.columns)}):")
            for i, col in enumerate(df.columns):
                print(f"  {i:2d}. '{col}'")
            
            print(f"\nTotal filas leídas: {len(df)}")
            
            # Buscar columnas por patrones si no tienen los nombres exactos
            column_mapping = {}
            
            # Buscar Total
            for col in df.columns:
                col_lower = str(col).lower()
                if 'total' in col_lower and 'base' not in col_lower:
                    column_mapping['Total'] = col
                elif 'valor' in col_lower and 'total' in col_lower:
                    column_mapping['Total'] = col
                elif 'monetario' in col_lower:
                    column_mapping['Total'] = col
            
            # Buscar IVA
            for col in df.columns:
                col_lower = str(col).lower()
                if 'iva' in col_lower and 'rete' not in col_lower and 'total' not in col_lower:
                    column_mapping['IVA'] = col
                elif 'impuesto' in col_lower and 'valor' in col_lower:
                    column_mapping['IVA'] = col
            
            # Buscar NIT Emisor
            for col in df.columns:
                col_lower = str(col).lower()
                if 'nit' in col_lower and 'emisor' in col_lower:
                    column_mapping['NIT Emisor'] = col
                elif 'documento' in col_lower and 'emisor' in col_lower:
                    column_mapping['NIT Emisor'] = col
            
            # Buscar Nombre Emisor
            for col in df.columns:
                col_lower = str(col).lower()
                if 'nombre' in col_lower and 'emisor' in col_lower:
                    column_mapping['Nombre Emisor'] = col
                elif 'razón' in col_lower and 'social' in col_lower:
                    column_mapping['Nombre Emisor'] = col
            
            # Buscar NIT Receptor
            for col in df.columns:
                col_lower = str(col).lower()
                if 'nit' in col_lower and 'receptor' in col_lower:
                    column_mapping['NIT Receptor'] = col
                elif 'documento' in col_lower and 'receptor' in col_lower:
                    column_mapping['NIT Receptor'] = col
            
            # Buscar Nombre Receptor
            for col in df.columns:
                col_lower = str(col).lower()
                if 'nombre' in col_lower and 'receptor' in col_lower:
                    column_mapping['Nombre Receptor'] = col
            
            print(f"\nMapeo de columnas encontrado: {column_mapping}")
            
            # Renombrar columnas a nombres estándar
            df = df.rename(columns=column_mapping)
            
            # Verificar que tenemos las columnas mínimas requeridas
            if 'Total' not in df.columns:
                # Intentar encontrar por índice (última columna numérica)
                numeric_cols = []
                for col in df.columns:
                    try:
                        # Verificar si la columna contiene números
                        sample = df[col].dropna().head(10)
                        if len(sample) > 0 and any(str(x).replace(',', '').replace('.', '').replace('-', '').isdigit() for x in sample):
                            numeric_cols.append(col)
                    except:
                        pass
                
                if numeric_cols:
                    # Usar la última columna numérica como Total
                    df = df.rename(columns={numeric_cols[-1]: 'Total'})
                    print(f"Usando '{numeric_cols[-1]}' como columna Total")
            
            # Filtrar solo Facturas electrónicas si existe la columna
            if 'Tipo de documento' in df.columns:
                original_count = len(df)
                mask = df['Tipo de documento'].astype(str).str.contains('Factura', case=False, na=False)
                df = df[mask]
                print(f"Facturas filtradas: {len(df)} de {original_count}")
            else:
                print("Advertencia: No se encontró columna 'Tipo de documento'")
            
            # Convertir columnas numéricas usando el método mejorado
            for col in df.columns:
                if col in ['Total', 'IVA', 'ICA', 'Rete IVA', 'Rete Renta', 'Rete ICA']:
                    try:
                        # Usar la función limpiar_numero para cada valor
                        df[col] = df[col].apply(self.limpiar_numero)
                        print(f"✓ Convertida columna {col} a numérico")
                        
                        # Mostrar muestra de valores para verificación
                        muestra = df[col].head(3).tolist()
                        print(f"  Muestra: {muestra}")
                    except Exception as e:
                        print(f"❌ Error convirtiendo columna {col}: {e}")
                        df[col] = 0
            
            # Si no se encontró columna IVA, calcularla si es posible
            if 'IVA' not in df.columns and 'Total' in df.columns:
                print("Advertencia: No se encontró columna IVA, se asumirá 0")
                df['IVA'] = 0
            
            return df
            
        except Exception as e:
            raise Exception(f"Error leyendo archivo: {str(e)}\nDetalle: {traceback.format_exc()}")
    
    def procesar_compras(self, df):
        """
        Procesa COMPRAS según especificaciones:
        - Débito (gasto) = Total - IVA
        - Débito (IVA) = IVA original
        - VALOR_BASE = IVA / 0.19 (redondeado al peso)
        - Todos los valores redondeados al peso más cercano
        """
        registros = []
        
        print(f"\nProcesando {len(df)} compras...")
        print(f"Columnas disponibles: {list(df.columns)}")
        
        # Verificar columnas requeridas
        required_cols = ['Total', 'IVA']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise Exception(f"Columnas faltantes para compras: {missing_cols}")
        
        for idx, row in df.iterrows():
            try:
                # Obtener valores - ya vienen como float del leer_archivo_dian
                total = float(row.get('Total', 0))
                iva = float(row.get('IVA', 0))
                
                # DEBUG: Mostrar primeros 3 registros
                if idx < 3:
                    print(f"\n📊 Registro {idx + 1}:")
                    print(f"   Total: {total:,.2f}")
                    print(f"   IVA: {iva:,.2f}")
                
                # Obtener NIT Emisor o usar valor por defecto
                if 'NIT Emisor' in df.columns:
                    nit = self.limpiar_nit(row.get('NIT Emisor', ''))
                else:
                    # Intentar encontrar columna con NIT
                    nit_cols = [col for col in df.columns if 'nit' in str(col).lower()]
                    nit = self.limpiar_nit(row[nit_cols[0]]) if nit_cols else ""
                
                # Obtener nombre o usar valor por defecto
                if 'Nombre Emisor' in df.columns:
                    obs = str(row.get('Nombre Emisor', f'Compra {idx+1}'))[:50]
                else:
                    obs = f"Compra {idx+1}"
                
                if total == 0 and iva == 0:
                    continue
                
                # Calcular valores
                valor_sin_iva = total - iva
                base_iva = iva / self.IVA_RATE if iva > 0 else 0
                
                # DEBUG: Mostrar cálculo
                if idx < 3:
                    print(f"   Valor sin IVA: {valor_sin_iva:,.2f}")
                    print(f"   Base IVA: {base_iva:,.2f}")
                
                # Redondear todos los valores al peso más cercano
                valor_sin_iva_entero = self.redondear_peso(valor_sin_iva)
                iva_entero = self.redondear_peso(iva)
                base_iva_entero = self.redondear_peso(base_iva)
                
                # DEBUG: Mostrar valores redondeados
                if idx < 3:
                    print(f"   Redondeados:")
                    print(f"   - Valor sin IVA: {valor_sin_iva_entero:,}")
                    print(f"   - IVA: {iva_entero:,}")
                    print(f"   - Base IVA: {base_iva_entero:,}")
                
                # Fila 1: Gasto (Débito) - VALORES ENTEROS
                registros.append({
                    'CUENTA': self.CUENTAS_COMPRAS['gasto'],
                    'CC': '',
                    'OBSERVACIONES': obs,
                    'DEBITO': valor_sin_iva_entero,
                    'CREDITO': None,
                    'VALOR_BASE': None,
                    'TERCERO': nit,
                    'H': None
                })
                
                # Fila 2: IVA descontable (Débito) - VALORES ENTEROS
                if iva > 0:
                    registros.append({
                        'CUENTA': self.CUENTAS_COMPRAS['iva_descontable'],
                        'CC': '',
                        'OBSERVACIONES': obs,
                        'DEBITO': iva_entero,
                        'CREDITO': None,
                        'VALOR_BASE': base_iva_entero,
                        'TERCERO': nit,
                        'H': 1
                    })
                    
            except Exception as e:
                print(f"❌ Error procesando fila {idx}: {e}")
                print(f"   Total: {row.get('Total', 'N/A')}")
                print(f"   IVA: {row.get('IVA', 'N/A')}")
                traceback.print_exc()
                continue
        
        print(f"\n✅ Registros generados: {len(registros)}")
        return pd.DataFrame(registros)
    
    def procesar_ventas(self, df):
        """
        Procesa VENTAS según especificaciones:
        - Crédito (ingresos) = Total - IVA
        - Crédito (IVA generado) = IVA original
        - Débito (IVA crédito) = Total factura (Base + IVA)
        - VALOR_BASE = IVA / 0.19 (redondeado al peso)
        - Todos los valores redondeados al peso más cercano
        """
        registros = []
        
        print(f"\nProcesando {len(df)} ventas...")
        print(f"Columnas disponibles: {list(df.columns)}")
        
        # Verificar columnas requeridas
        required_cols = ['Total', 'IVA']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise Exception(f"Columnas faltantes para ventas: {missing_cols}")
        
        for idx, row in df.iterrows():
            try:
                # Obtener valores - ya vienen como float del leer_archivo_dian
                total = float(row.get('Total', 0))
                iva = float(row.get('IVA', 0))
                
                # DEBUG: Mostrar primeros 3 registros
                if idx < 3:
                    print(f"\n📊 Registro {idx + 1}:")
                    print(f"   Total: {total:,.2f}")
                    print(f"   IVA: {iva:,.2f}")
                
                # Obtener NIT Receptor o usar valor por defecto
                if 'NIT Receptor' in df.columns:
                    nit = self.limpiar_nit(row.get('NIT Receptor', ''))
                else:
                    # Intentar encontrar columna con NIT
                    nit_cols = [col for col in df.columns if 'nit' in str(col).lower()]
                    nit = self.limpiar_nit(row[nit_cols[0]]) if nit_cols else ""
                
                # Obtener nombre o usar valor por defecto
                if 'Nombre Receptor' in df.columns:
                    obs = str(row.get('Nombre Receptor', f'Venta {idx+1}'))[:50]
                else:
                    obs = f"Venta {idx+1}"
                
                if total == 0 and iva == 0:
                    continue
                
                # Calcular valores
                valor_sin_iva = total - iva
                base_iva = iva / self.IVA_RATE if iva > 0 else 0
                
                # DEBUG: Mostrar cálculo
                if idx < 3:
                    print(f"   Valor sin IVA: {valor_sin_iva:,.2f}")
                    print(f"   Base IVA: {base_iva:,.2f}")
                
                # Redondear todos los valores al peso más cercano
                valor_sin_iva_entero = self.redondear_peso(valor_sin_iva)
                iva_entero = self.redondear_peso(iva)
                base_iva_entero = self.redondear_peso(base_iva)
                total_entero = self.redondear_peso(total)
                
                # DEBUG: Mostrar valores redondeados
                if idx < 3:
                    print(f"   Redondeados:")
                    print(f"   - Valor sin IVA: {valor_sin_iva_entero:,}")
                    print(f"   - IVA: {iva_entero:,}")
                    print(f"   - Base IVA: {base_iva_entero:,}")
                    print(f"   - Total: {total_entero:,}")
                
                # Fila 1: Ingresos (Crédito) - Cuenta 41 - VALORES ENTEROS
                registros.append({
                    'CUENTA': self.CUENTAS_VENTAS['ingresos'],
                    'CC': '',
                    'OBSERVACIONES': obs,
                    'DEBITO': None,
                    'CREDITO': valor_sin_iva_entero,
                    'VALOR_BASE': None,
                    'TERCERO': nit,
                    'H': None
                })
                
                # Fila 2: IVA Generado (Crédito) - Cuenta 24080101 - VALORES ENTEROS
                if iva > 0:
                    registros.append({
                        'CUENTA': self.CUENTAS_VENTAS['iva_generado'],
                        'CC': '',
                        'OBSERVACIONES': obs,
                        'DEBITO': None,
                        'CREDITO': iva_entero,
                        'VALOR_BASE': base_iva_entero,
                        'TERCERO': nit,
                        'H': 1
                    })
                    
                    # Fila 3: IVA Débito - Cuenta 13050501 - VALORES ENTEROS
                    registros.append({
                        'CUENTA': self.CUENTAS_VENTAS['iva_credito'],
                        'CC': '',
                        'OBSERVACIONES': obs,
                        'DEBITO': total_entero,  # Total factura (Base + IVA)
                        'CREDITO': None,
                        'VALOR_BASE': None,
                        'TERCERO': nit,
                        'H': None
                    })
                    
            except Exception as e:
                print(f"❌ Error procesando fila {idx}: {e}")
                print(f"   Total: {row.get('Total', 'N/A')}")
                print(f"   IVA: {row.get('IVA', 'N/A')}")
                traceback.print_exc()
                continue
        
        print(f"\n✅ Registros generados: {len(registros)}")
        return pd.DataFrame(registros)


class AplicacionDIAN:
    """Interfaz gráfica"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("DIAN → Siigo | Conversor Contable v3.1")
        self.root.geometry("1000x800")
        
        # Paleta de colores pasteles en rosas
        self.COLORES = {
            'fondo_principal': '#FFF0F5',
            'fondo_secundario': '#FFE6F2',
            'fondo_frame': '#FFFFFF',
            'titulo_principal': '#C71585',
            'texto_principal': '#8B0058',
            'texto_secundario': '#A0527A',
            'boton_principal': '#FFB6C1',
            'boton_secundario': '#FFC0CB',
            'boton_accion': '#FF69B4',
            'boton_exito': '#FF1493',
            'boton_peligro': '#DB7093',
            'barra_progreso': '#FFC0CB',
            'log_fondo': '#2c3e50',
            'log_texto': '#FFB6C1',
            'borde': '#FFB6C1'
        }
        
        self.root.configure(bg=self.COLORES['fondo_principal'])
        
        self.procesador = ProcesadorContableDIAN()
        self.archivo_actual = None
        self.df_resultado = None
        
        self.crear_widgets()
    
    def crear_widgets(self):
        main_frame = tk.Frame(self.root, bg=self.COLORES['fondo_principal'], padx=30, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Título
        tk.Label(main_frame, text="Conversor DIAN a Siigo", 
                font=('Helvetica', 26, 'bold'), 
                bg=self.COLORES['fondo_principal'], 
                fg=self.COLORES['titulo_principal']).pack(pady=(0, 5))
        
        tk.Label(main_frame, text="Procesa archivos de compras y ventas descargados de la DIAN", 
                font=('Helvetica', 12), 
                bg=self.COLORES['fondo_principal'], 
                fg=self.COLORES['texto_secundario']).pack(pady=(0, 20))
        
        # Frame de selección
        frame_archivo = tk.LabelFrame(main_frame, text=" 1. Seleccionar Archivo de la DIAN ", 
                                     bg=self.COLORES['fondo_frame'], 
                                     fg=self.COLORES['texto_principal'],
                                     font=('Helvetica', 12, 'bold'), 
                                     padx=15, pady=15,
                                     relief=tk.RIDGE,
                                     borderwidth=2)
        frame_archivo.pack(fill=tk.X, pady=10)
        
        self.entry_ruta = tk.Entry(frame_archivo, font=('Helvetica', 11), 
                                  width=65, relief=tk.SOLID, bd=2,
                                  bg='white', fg=self.COLORES['texto_principal'])
        self.entry_ruta.pack(side=tk.LEFT, padx=(0, 10), ipady=5)
        
        tk.Button(frame_archivo, text="📁 Buscar Archivo", 
                 command=self.seleccionar_archivo,
                 bg=self.COLORES['boton_principal'], 
                 fg=self.COLORES['texto_principal'],
                 font=('Helvetica', 11, 'bold'),
                 relief=tk.RAISED, 
                 padx=20, pady=6,
                 cursor='hand2',
                 activebackground=self.COLORES['boton_accion'],
                 activeforeground='white').pack(side=tk.LEFT)
        
        # Frame de tipo
        frame_tipo = tk.LabelFrame(main_frame, text=" 2. Tipo de Documento ", 
                                  bg=self.COLORES['fondo_frame'],
                                  fg=self.COLORES['texto_principal'],
                                  font=('Helvetica', 12, 'bold'), 
                                  padx=15, pady=15,
                                  relief=tk.RIDGE,
                                  borderwidth=2)
        frame_tipo.pack(fill=tk.X, pady=10)
        
        self.tipo_var = tk.StringVar(value="auto")
        
        tipos = [
            ("🔍 Detectar automáticamente (recomendado)", "auto"),
            ("🛒 Compras / Recibidos (usará NIT Emisor)", "compras"),
            ("💰 Ventas / Enviados (usará NIT Receptor)", "ventas")
        ]
        
        for texto, valor in tipos:
            tk.Radiobutton(frame_tipo, text=texto, variable=self.tipo_var, 
                          value=valor, 
                          bg=self.COLORES['fondo_frame'], 
                          fg=self.COLORES['texto_principal'],
                          font=('Helvetica', 11),
                          selectcolor=self.COLORES['boton_principal'],
                          activebackground=self.COLORES['fondo_frame'],
                          activeforeground=self.COLORES['texto_principal']).pack(anchor=tk.W, pady=4)
        
        # Botón procesar
        tk.Button(main_frame, text="⚡ PROCESAR ARCHIVO", 
                 command=self.procesar_archivo,
                 bg=self.COLORES['boton_accion'], 
                 fg='white',
                 font=('Helvetica', 16, 'bold'),
                 relief=tk.RAISED, 
                 padx=40, pady=15,
                 cursor='hand2',
                 activebackground=self.COLORES['boton_exito'],
                 activeforeground='white').pack(pady=20)
        
        # Barra de progreso
        style = ttk.Style()
        style.theme_use('default')
        style.configure("TProgressbar",
                       thickness=20,
                       background=self.COLORES['barra_progreso'],
                       troughcolor=self.COLORES['fondo_secundario'])
        
        self.progress = ttk.Progressbar(main_frame, orient=tk.HORIZONTAL, 
                                       length=500, mode='determinate',
                                       style="TProgressbar")
        self.progress.pack(pady=10)
        
        # Frame de resultados
        self.frame_resultados = tk.LabelFrame(main_frame, text=" 3. Resultados y Log ", 
                                             bg=self.COLORES['fondo_frame'],
                                             fg=self.COLORES['texto_principal'],
                                             font=('Helvetica', 12, 'bold'), 
                                             padx=15, pady=15,
                                             relief=tk.RIDGE,
                                             borderwidth=2)
        self.frame_resultados.pack(fill=tk.BOTH, expand=True, pady=10)
        
        self.lbl_estado = tk.Label(self.frame_resultados, 
                                  text="Esperando archivo...", 
                                  bg=self.COLORES['fondo_frame'], 
                                  fg=self.COLORES['texto_secundario'], 
                                  font=('Helvetica', 12))
        self.lbl_estado.pack(pady=5)
        
        # Botones
        self.frame_botones = tk.Frame(self.frame_resultados, bg=self.COLORES['fondo_frame'])
        self.frame_botones.pack(pady=5)
        
        self.btn_ver = tk.Button(self.frame_botones, text="👁 Ver Vista Previa", 
                                command=self.ver_preview, state=tk.DISABLED,
                                bg=self.COLORES['boton_peligro'], 
                                fg='white',
                                font=('Helvetica', 11, 'bold'),
                                relief=tk.RAISED, 
                                padx=15, pady=8,
                                cursor='hand2',
                                activebackground='#C71585',
                                activeforeground='white',
                                disabledforeground='white')
        self.btn_ver.pack(side=tk.LEFT, padx=5)
        
        self.btn_excel = tk.Button(self.frame_botones, text="💾 Descargar Excel", 
                                  command=self.guardar_excel, state=tk.DISABLED,
                                  bg=self.COLORES['boton_peligro'], 
                                  fg='white',
                                  font=('Helvetica', 11, 'bold'),
                                  relief=tk.RAISED, 
                                  padx=15, pady=8,
                                  cursor='hand2',
                                  activebackground='#C71585',
                                  activeforeground='white',
                                  disabledforeground='white')
        self.btn_excel.pack(side=tk.LEFT, padx=5)
        
        self.btn_query = tk.Button(self.frame_botones, text="📋 Power Query", 
                                  command=self.mostrar_power_query, state=tk.DISABLED,
                                  bg=self.COLORES['boton_peligro'], 
                                  fg='white',
                                  font=('Helvetica', 11, 'bold'),
                                  relief=tk.RAISED, 
                                  padx=15, pady=8,
                                  cursor='hand2',
                                  activebackground='#C71585',
                                  activeforeground='white',
                                  disabledforeground='white')
        self.btn_query.pack(side=tk.LEFT, padx=5)
        
        # Resumen
        self.lbl_resumen = tk.Label(self.frame_resultados, text="", 
                                   bg=self.COLORES['fondo_frame'], 
                                   fg=self.COLORES['texto_principal'],
                                   font=('Helvetica', 11, 'bold'), 
                                   justify=tk.LEFT)
        self.lbl_resumen.pack(pady=5)
        
        # Log
        self.txt_log = scrolledtext.ScrolledText(self.frame_resultados, 
                                                height=12, width=90,
                                                font=('Consolas', 10),
                                                bg=self.COLORES['log_fondo'], 
                                                fg=self.COLORES['log_texto'],
                                                relief=tk.SUNKEN,
                                                borderwidth=2)
        self.txt_log.pack(fill=tk.BOTH, expand=True, pady=10)
        self.txt_log.insert(tk.END, "Log de procesamiento iniciado...\n")
        self.txt_log.config(state=tk.DISABLED)
    
    def log(self, mensaje):
        """Agrega mensaje al log"""
        self.txt_log.config(state=tk.NORMAL)
        self.txt_log.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {mensaje}\n")
        self.txt_log.see(tk.END)
        self.txt_log.config(state=tk.DISABLED)
        self.root.update()
    
    def seleccionar_archivo(self):
        """Abre diálogo para seleccionar archivo"""
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo de la DIAN",
            filetypes=[
                ("Archivos Excel", "*.xlsx *.xls"),
                ("Archivos CSV", "*.csv"),
                ("Todos los archivos", "*.*")
            ]
        )
        if archivo:
            self.archivo_actual = archivo
            self.entry_ruta.delete(0, tk.END)
            self.entry_ruta.insert(0, archivo)
            nombre = Path(archivo).name
            self.lbl_estado.config(text=f"Archivo seleccionado: {nombre}", fg=self.COLORES['boton_accion'])
            self.log(f"Archivo seleccionado: {nombre}")
            
            # Detectar tipo por nombre
            nombre_lower = nombre.lower()
            if 'recibido' in nombre_lower:
                self.tipo_var.set('compras')
                self.log("Tipo sugerido: Compras")
            elif 'enviado' in nombre_lower:
                self.tipo_var.set('ventas')
                self.log("Tipo sugerido: Ventas")
    
    def procesar_archivo(self):
        """Procesa el archivo seleccionado"""
        if not self.archivo_actual:
            messagebox.showwarning("Atención", "Por favor selecciona un archivo primero.")
            return
        
        # Limpiar log
        self.txt_log.config(state=tk.NORMAL)
        self.txt_log.delete(1.0, tk.END)
        self.txt_log.config(state=tk.DISABLED)
        
        try:
            self.progress['value'] = 10
            self.root.update()
            
            # Leer archivo
            self.log("Leyendo archivo...")
            df = self.procesador.leer_archivo_dian(self.archivo_actual)
            self.log(f"Filas leídas: {len(df)}")
            
            if len(df) == 0:
                raise Exception("No se encontraron facturas en el archivo.")
            
            self.progress['value'] = 30
            
            # Mostrar columnas detectadas
            self.log(f"Columnas detectadas: {', '.join(list(df.columns))}")
            
            # Determinar tipo
            tipo = self.tipo_var.get()
            if tipo == "auto":
                columnas_str = ' '.join([str(c).upper() for c in df.columns])
                if 'NIT EMISOR' in columnas_str:
                    tipo = "compras"
                    self.log("Tipo detectado: Compras (por NIT Emisor)")
                elif 'NIT RECEPTOR' in columnas_str:
                    tipo = "ventas"
                    self.log("Tipo detectado: Ventas (por NIT Receptor)")
                else:
                    tipo = "compras"
                    self.log("Tipo por defecto: Compras")
            
            self.progress['value'] = 50
            
            # Verificar que existan las columnas necesarias para el tipo seleccionado
            if tipo == "compras":
                if 'NIT Emisor' not in df.columns:
                    self.log("Advertencia: No se encontró 'NIT Emisor', usando valor por defecto")
                if 'Nombre Emisor' not in df.columns:
                    self.log("Advertencia: No se encontró 'Nombre Emisor', usando valor por defecto")
            else:
                if 'NIT Receptor' not in df.columns:
                    self.log("Advertencia: No se encontró 'NIT Receptor', usando valor por defecto")
                if 'Nombre Receptor' not in df.columns:
                    self.log("Advertencia: No se encontró 'Nombre Receptor', usando valor por defecto")
            
            # Verificar columnas comunes
            for col in ['Total', 'IVA']:
                if col not in df.columns:
                    raise Exception(f"No se encontró la columna '{col}' en el archivo")
            
            self.log(f"Columnas verificadas: Total, IVA presentes")
            self.progress['value'] = 70
            
            # Procesar según tipo
            self.log(f"Procesando como {tipo}...")
            if tipo == "compras":
                self.df_resultado = self.procesador.procesar_compras(df)
                tipo_nombre = "Compras/Recibidos"
            else:
                self.df_resultado = self.procesador.procesar_ventas(df)
                tipo_nombre = "Ventas/Enviados"
            
            self.progress['value'] = 100
            
            # Verificar resultado
            if len(self.df_resultado) == 0:
                self.log("⚠️ ERROR: No se generaron registros")
                messagebox.showerror("Error", 
                    "No se generaron registros.\n"
                    "Verifica que las facturas tengan valores en Total e IVA.")
                self.progress['value'] = 0
                return
            
            # Éxito
            self.log(f"✅ ÉXITO: {len(self.df_resultado)} filas generadas")
            self.lbl_estado.config(text=f"✅ Completado: {tipo_nombre}", fg=self.COLORES['boton_exito'])
            self.lbl_resumen.config(text=f"Tipo: {tipo_nombre}\n"
                                        f"Filas generadas: {len(self.df_resultado)}\n"
                                        f"Facturas procesadas: {len(df)}\n"
                                        f"Valores redondeados al peso más cercano")
            
            # Habilitar botones
            self.btn_ver.config(state=tk.NORMAL, bg=self.COLORES['boton_accion'], fg='white')
            self.btn_excel.config(state=tk.NORMAL)
            self.btn_query.config(state=tk.NORMAL)
            
            messagebox.showinfo("Éxito", 
                f"Procesamiento completado.\n\n"
                f"Tipo: {tipo_nombre}\n"
                f"Facturas: {len(df)}\n"
                f"Registros Siigo: {len(self.df_resultado)}\n\n"
                f"NOTAS:\n"
                f"✓ Todos los valores redondeados al peso más cercano\n"
                f"✓ VALOR_BASE calculado como IVA/0.19 (sin decimales)\n"
                f"✓ Formato colombiano: 200.000,00")
            
        except Exception as e:
            self.progress['value'] = 0
            error_msg = str(e)
            self.log(f"❌ ERROR: {error_msg}")
            self.log(traceback.format_exc())
            messagebox.showerror("Error", f"Error al procesar:\n\n{error_msg}")
    
    def formato_display(self, valor):
        """Formatea un valor numérico para mostrar en la vista previa"""
        if pd.isna(valor) or valor is None:
            return ""
        if isinstance(valor, (int, float)):
            # Redondear al entero más cercano
            valor_entero = int(round(float(valor), 0))
            # Formato colombiano: 200.000,00
            return f"{valor_entero:,}".replace(",", ".") + ",00"
        return str(valor)
    
    def ver_preview(self):
        """Muestra ventana con vista previa"""
        if self.df_resultado is None or len(self.df_resultado) == 0:
            return
        
        ventana = tk.Toplevel(self.root)
        ventana.title("Vista Previa - Datos para Siigo")
        ventana.geometry("1100x600")
        ventana.configure(bg=self.COLORES['fondo_principal'])
        
        frame = tk.Frame(ventana, padx=10, pady=10, bg=self.COLORES['fondo_principal'])
        frame.pack(fill=tk.BOTH, expand=True)
        
        columnas = list(self.df_resultado.columns)
        tree = ttk.Treeview(frame, columns=columnas, show='headings', height=20)
        
        # Configurar estilo para el treeview
        style = ttk.Style()
        style.configure("Treeview",
                      background=self.COLORES['fondo_frame'],
                      foreground=self.COLORES['texto_principal'],
                      fieldbackground=self.COLORES['fondo_frame'])
        style.configure("Treeview.Heading",
                      background=self.COLORES['boton_principal'],
                      foreground=self.COLORES['texto_principal'],
                      font=('Helvetica', 10, 'bold'))
        
        for col in columnas:
            tree.heading(col, text=col)
            ancho = 150 if col in ['OBSERVACIONES', 'TERCERO'] else 100
            tree.column(col, width=ancho, anchor='center')
        
        # Insertar datos con formato de display
        for idx, row in self.df_resultado.head(100).iterrows():
            valores_display = []
            for col in columnas:
                valor = row[col]
                if col in ['DEBITO', 'CREDITO', 'VALOR_BASE']:
                    valores_display.append(self.formato_display(valor))
                elif col == 'H':
                    valores_display.append(str(valor) if pd.notna(valor) and valor != '' else '')
                else:
                    valores_display.append(str(valor) if pd.notna(valor) and valor != '' else '')
            tree.insert('', tk.END, values=valores_display)
        
        scrollbar_y = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        scrollbar_x = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        tree.grid(row=0, column=0, sticky='nsew')
        scrollbar_y.grid(row=0, column=1, sticky='ns')
        scrollbar_x.grid(row=1, column=0, sticky='ew')
        
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
        
        tk.Label(ventana, text=f"Mostrando {min(100, len(self.df_resultado))} de {len(self.df_resultado)} filas", 
                fg=self.COLORES['texto_secundario'], 
                bg=self.COLORES['fondo_principal'],
                font=('Helvetica', 9)).pack(pady=5)
    
    def guardar_excel(self):
        """Guarda el resultado en Excel con formato colombiano EXACTO"""
        if self.df_resultado is None:
            return
        
        tipo = self.tipo_var.get()
        prefijo = "Compras" if tipo == "compras" else "Ventas"
        
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"{prefijo}_Siigo_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
        if archivo:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import numbers
                
                # Crear un nuevo workbook
                wb = Workbook()
                ws = wb.active
                ws.title = 'Siigo'
                
                # Escribir encabezados
                for col_idx, col_name in enumerate(self.df_resultado.columns, start=1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                
                # Escribir datos - VALORES COMO ENTEROS (redondeados al peso)
                for row_idx, row_data in enumerate(self.df_resultado.itertuples(index=False), start=2):
                    for col_idx, (col_name, valor) in enumerate(zip(self.df_resultado.columns, row_data), start=1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        
                        # Escribir valor según el tipo
                        if col_name in ['DEBITO', 'CREDITO', 'VALOR_BASE']:
                            if pd.notna(valor) and valor is not None:
                                # Asegurar que sea entero
                                cell.value = int(round(float(valor), 0))
                                # Formato colombiano EXACTO: 200.000,00
                                cell.number_format = '#.##0,00'
                            else:
                                cell.value = None
                        elif col_name == 'H':
                            if pd.notna(valor) and valor is not None and valor != '':
                                cell.value = int(valor)
                            else:
                                cell.value = None
                        else:
                            cell.value = str(valor) if pd.notna(valor) and valor != '' else ''
                
                # Ajustar anchos de columna
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
                
                # Guardar archivo
                wb.save(archivo)
                
                self.log(f"✅ Excel guardado: {archivo}")
                self.log("✅ Valores guardados como ENTEROS (redondeados al peso)")
                self.log("✅ Formato colombiano EXACTO: #.##0,00")
                
                if messagebox.askyesno("Éxito", 
                    "Archivo guardado exitosamente.\n\n"
                    "✅ Todos los valores redondeados al peso más cercano\n"
                    "✅ Formato colombiano EXACTO: 200.000,00\n"
                    "✅ Listo para usar en Siigo\n\n"
                    "¿Deseas abrir el archivo ahora?"):
                    os.startfile(archivo)
                    
            except Exception as e:
                self.log(f"❌ Error guardando: {str(e)}")
                messagebox.showerror("Error", f"No se pudo guardar:\n{str(e)}")
    
    def mostrar_power_query(self):
        """Muestra código Power Query con valores enteros"""
        if self.df_resultado is None:
            return
        
        # Crear copia del dataframe para formatear valores en el código M
        df_display = self.df_resultado.copy()
        
        # Generar código M
        filas = []
        for idx, row in df_display.iterrows():
            valores = []
            for col_name, v in zip(df_display.columns, row.values):
                if col_name in ['DEBITO', 'CREDITO', 'VALOR_BASE']:
                    if pd.isna(v) or v is None:
                        valores.append('null')
                    else:
                        # Convertir a entero (ya redondeado)
                        valores.append(str(int(round(float(v), 0))))
                elif col_name == 'H':
                    if pd.isna(v) or v is None or v == '':
                        valores.append('null')
                    else:
                        valores.append(str(int(v)))
                else:
                    if pd.isna(v) or v == '':
                        valores.append('null')
                    else:
                        valores.append(f'"{str(v)}"')
            filas.append(f"    {{ {', '.join(valores)} }}")
        
        datos = ",\n".join(filas)
        headers = ", ".join([f'"{col}"' for col in df_display.columns])
        
        codigo_m = f"""let
    Origen = #table(
        {{ {headers} }},
        {{
{datos}
        }}
    ),
    TipoCambiado = Table.TransformColumnTypes(Origen,{{
        {{"CUENTA", type text}}, 
        {{"CC", type text}}, 
        {{"OBSERVACIONES", type text}}, 
        {{"DEBITO", Int64.Type}}, 
        {{"CREDITO", Int64.Type}}, 
        {{"VALOR_BASE", Int64.Type}},
        {{"TERCERO", type text}}, 
        {{"H", Int64.Type}}
    }}),
    Limpieza = Table.ReplaceValue(TipoCambiado,null,null,Replacer.ReplaceValue,
        {{"DEBITO", "CREDITO", "H", "VALOR_BASE"}}),
    Filtrado = Table.SelectRows(Limpieza, each ([CUENTA] <> null))
in
    Filtrado"""
        
        ventana = tk.Toplevel(self.root)
        ventana.title("Código Power Query (M)")
        ventana.geometry("900x700")
        ventana.configure(bg=self.COLORES['fondo_principal'])
        
        tk.Label(ventana, text="Copia este código en Excel (Datos > Obtener datos > Editor avanzado)", 
                fg=self.COLORES['texto_principal'], 
                bg=self.COLORES['fondo_principal'],
                font=('Helvetica', 11, 'bold')).pack(pady=10)
        
        texto = scrolledtext.ScrolledText(ventana, wrap=tk.WORD, 
                                         font=('Consolas', 10), height=30,
                                         bg=self.COLORES['log_fondo'], 
                                         fg=self.COLORES['log_texto'])
        texto.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        texto.insert(tk.END, codigo_m)
        
        def copiar():
            self.root.clipboard_clear()
            self.root.clipboard_append(codigo_m)
            messagebox.showinfo("Copiado", "Código copiado al portapapeles")
        
        tk.Button(ventana, text="📋 Copiar al portapapeles", 
                 command=copiar,
                 bg=self.COLORES['boton_exito'], 
                 fg='white',
                 font=('Helvetica', 12, 'bold'),
                 relief=tk.RAISED, 
                 padx=20, pady=10,
                 cursor='hand2',
                 activebackground='#FF0066',
                 activeforeground='white').pack(pady=10)


if __name__ == "__main__":
    # Instalar dependencias si faltan
    try:
        import pandas
        import openpyxl
    except ImportError:
        print("Instalando dependencias...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", 
                             "pandas", "openpyxl", "-q"])
        print("Dependencias instaladas. Reiniciando...")
        os.execv(sys.executable, ['python'] + sys.argv)
    
    # Iniciar
    root = tk.Tk()
    app = AplicacionDIAN(root)
    root.mainloop()